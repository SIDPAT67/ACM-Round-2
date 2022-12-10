
import json
from openpyxl import load_workbook
wb = load_workbook(filename = 'excel_2.xlsx')

sheet = wb['One']
books = {}

for row in sheet.iter_rows(values_only = True):
    book_id = row[0]
    book ={
        "title": row[1],
        "author":row[2],
    }
   
books[book_id] = book

print(json.dumps(books, indent = 3))

#wb.save("excel_2.xlsx")
